import re
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date
import shutil

# ======================== Função inicial ===========================
def separar_em_paragrafo(caminho_arquivo):
    paragrafos = []

    # ler arquivo  'cadastro_texto.txt'
    with open(caminho_arquivo, 'r', encoding='utf8') as arquivo: # with (abre e fecha)
        texto = arquivo.read()

    for linha in texto.split("\n"):
        if linha:
            paragrafos.append(linha)
    return paragrafos

# ======================== Funções de filtro ===========================
# Filtrar portaria
def filtrar_portaria(portaria, resolucao):
    pre_filtro_portaria = re.findall(r'\bPor\s?-?ta\s?-?ria.*', str(*portaria), flags=re.I)
    _filtrar_portaria_ = str(*pre_filtro_portaria)
    _filtrar_resolucao_ = str(*resolucao)
    print(_filtrar_portaria_)

    portaria.clear()
    resolucao.clear()

    # Funções auxiliares
    _filtrar_portaria_ = eliminar_virgula_no_texto_portaria(_filtrar_portaria_)
    _filtrar_resolucao_ = eliminar_virgula_no_texto_resolucao(_filtrar_resolucao_)

    if _filtrar_portaria_:
        portaria.append(_filtrar_portaria_)

    if _filtrar_resolucao_:
        portaria.append(_filtrar_resolucao_)

# Filtrar matricula
def filtrar_matricula(matricula):
    _filtrar_matricula_1 = re.sub(r'[a-zA-ZÀ-ú:,]', '' ,str(*matricula), flags=re.I)
    matricula.clear()
    _filtrar_matricula_2 = re.findall(r'\d{2}/\d{3}.\d{3}-\d', _filtrar_matricula_1)

    matricula.append(str(*_filtrar_matricula_2))

# Filtrar profissão
def filtrar_profissao(profissao):
    _filtrar_profissao_1 = re.findall(r'.+ ma-?tr', str(*profissao), flags=re.I)

    _filtrar_profissao_2 = re.sub(r'ma-?tr', '', str(*_filtrar_profissao_1), flags=re.I)

    profissao_string = _filtrar_profissao_2

    profissao.clear()
    variavel_auxiliar = ''
    for i in profissao_string:
        if i == ',':
            ...
        else:
            variavel_auxiliar += i

    profissao.append(variavel_auxiliar.strip())

# Filtrar nome
def filtrar_nome(nome):
        _filtrar_nome_ = str(*nome)
        nome.clear()

        variavel_auxiliar = ''
        iniciar = False
        for n1 in _filtrar_nome_:

            if n1.isupper():
                iniciar = True

            if n1 == ',':
                break

            if iniciar:
                if n1 == '-':
                    continue
                else:
                    variavel_auxiliar += n1

        nome.append(variavel_auxiliar)

# ======================== Funções auxiliares ==========================
def eliminar_virgula_no_texto_portaria(texto_string):
    eliminar_virgula_texto_string = ''

    for letra in texto_string:
        if letra == ',':
            break

        eliminar_virgula_texto_string += letra

    return eliminar_virgula_texto_string.strip()

def eliminar_virgula_no_texto_resolucao(texto_string):
    eliminar_virgula_texto_string = ''

    for letra in texto_string:
        if letra == ',':
            continue
        else:
            eliminar_virgula_texto_string += letra

    return eliminar_virgula_texto_string.strip()

# def unir_lista_em_um_texto(valor):
#         lista = []
#         variavel_auxiliar = ''
#         for i in valor:
#             variavel_auxiliar += i
#             variavel_auxiliar += ' '
#         lista = variavel_auxiliar.strip()
#         return lista

# ======================== Iniciar Programa ===========================
def cadastrar_pessoas():

    # caminho da pasta raiz
    PASTA_RAIZ = Path(__file__).parent
    PASTA_NOVA = PASTA_RAIZ / 'CADASTROS' / 'cadastro_texto.txt'

    # apagar pasta (delete)
    PASTA_PDF = PASTA_RAIZ / 'PDF_EM_TXT'
    shutil.rmtree(PASTA_PDF, ignore_errors=True)

    # finalizar caso pasta não exista
    if not os.path.exists(f'{PASTA_NOVA}'):
        return

    print(PASTA_NOVA)

    # Função inicial
    minha_lista = separar_em_paragrafo(PASTA_NOVA)

    # lista de dados dos clientes
    lista_de_clientes = []

    # ================== Logica de seleção de palavras (regras de negocio )=====================
    for texto in minha_lista:
        # =============================== Filtrar nome =====================================
        nome = re.findall(r'\bprov.+[\w]+', texto)
        filtrar_nome(nome)
        nome_string = str(*nome)
        print( nome_string)

        # ============================= Filtrar inatividade ===============================
        inatividade = re.findall(r'ina-?t.*-?e',texto, flags=re.I)
        if inatividade:
            inatividade = ['inativo']
            inatividade_string = str(*inatividade)
            print(inatividade_string)
        else:
            inatividade = ['NÃO inativo']
            inatividade_string = str(*inatividade)
            print(inatividade_string)

        # =============================== Filtrar profissão ================================
        profissao = re.findall(r',.?[0-9A-ZÀ-ú,-].+',texto)
        filtrar_profissao(profissao)
        profissao_string = str(*profissao)
        print(profissao_string)

        # ============================== Filtrar matricula =================================
        matricula = re.findall(r'ma-?tr.+Apo', texto, flags=re.I)
        filtrar_matricula(matricula)
        matricula_string =  str(*matricula)
        print(matricula_string)

        # ============================= Filtrar aposentando ================================
        aposentado = re.findall(r'\bapo.*do\b',texto, flags=re.I)
        aposentada = re.findall(r'\bapo.*da\b',texto, flags=re.I)
        if aposentada:
            aposentado_string = 'SIM aposentado(a)'
            print(aposentado_string)
        elif aposentado:
            aposentado_string = 'SIM aposentado(a)'
            print(aposentado_string)
        else:
            aposentado_string = 'NÃO aposentado(a)'
            print(aposentado_string)

        # ============================== Filtrar portaria =================================
        portaria = re.findall(r'\bapo.*\bPor\s?-?ta\s?-?ria.*',texto, flags=re.I)
        resolucao = re.findall(r'\bre\s?-?so\s?-?lu\s?-?ção.*\d{4}\b\s?,',texto, flags=re.I)
        filtrar_portaria(portaria, resolucao)
        portaria_string = str(*portaria)
        print(portaria_string) # Portaria

        # =============================== Filtrar processo ================================
        processo = re.findall(r'\bpro\s?-?ce\s?-?s\s?-?so\b.*',texto, flags=re.I)
        processo_string = str(*processo)
        print(processo_string)

        # =========================== Lista de clientes ===========================
        # dados individual dos clientes
        cliente = [ nome_string, inatividade_string, profissao_string, matricula_string,
                            aposentado_string, portaria_string, processo_string]

        lista_de_clientes.append(cliente)
        print('\n===========================================================\n')


    # =========================== Gerar arquivo EXCEL ===========================
    # caminho do arquivo
    hoje = date.today().strftime('%d.%m.%Y') # data de hoje (dd/mm/aaaa)
    CAMINHO_EXCEL = PASTA_RAIZ / 'CADASTROS' / 'cadastro.xlsx'

    workbook = Workbook() # criar arquivo vazio do Excel

    # Nome para a planilha
    sheet_name = f'RJ {hoje}'

    # Criamos a planilha
    workbook.create_sheet(sheet_name, 0) # alterar ordem (nome, posição)

    # Selecionou a planilha (worksheet: Worksheet = tipagem)
    worksheet: Worksheet = workbook[sheet_name] # planinha ativa ('Minha planilha')


    # Criando os cabeçalhos
    worksheet.cell(1, 1, 'NOME')    # (linha, coluna, valor)
    worksheet.cell(1, 2, 'SITUAÇÃO ATUAL')   # (linha, coluna, valor)
    worksheet.cell(1, 3, 'PROFISSÃO')    # (linha, coluna, valor)
    worksheet.cell(1, 4, 'MATRICULA')    # (linha, coluna, valor)
    worksheet.cell(1, 5, 'APOSENTADORIA')    # (linha, coluna, valor)
    worksheet.cell(1, 6, 'PORTARIA OU RESOLUÇÃO')    # (linha, coluna, valor)
    worksheet.cell(1, 7, 'NÚMERO DO PROCESSO')    # (linha, coluna, valor)

    lista_com_todos_os_clientes = list(lista_de_clientes)

    # adicionar lista de objetos
    for cliente in lista_com_todos_os_clientes:
        worksheet.append(cliente) # adicionar como lista
        print(cliente)
    # ===============================================================================

    workbook.save(CAMINHO_EXCEL) # salvar arquivo

# mensagem final
print('FIM Etapa_03_cadastro...')

if __name__ == "__main__":
    cadastrar_pessoas()
